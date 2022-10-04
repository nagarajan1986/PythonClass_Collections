


#Datadriven testing:  get data from excel sheet:


import openpyxl         #this package for read and write xls


#Read the data from excel sheet:        #File===>workbook====>sheet===>row====>cell

file="D:\Demofiles\python_demo.xlsx"        #excel file path need to save in one variable

workbook=openpyxl.load_workbook(file)       #first, load the workbook using openpyxl module
sheet=workbook["Sheet1"]                    #get sheet from workbook (or) use workbook.active if we have only one sheet in excel

rows=sheet.max_row                  #count number of rows in excel sheet
cols=sheet.max_column            #count number of columns in excel sheet

for r in range(1,rows+1):               #Read all the data from the excel sheet
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value, end='  ')      #get cell data using sheet.cell().value function
    print()


#Write the data into excel sheet:
file="D:\Demofiles\python_demo1.xlsx"        #excel file path need to save in one variable

workbook=openpyxl.load_workbook(file)       #first, load the workbook using openpyxl module
sheet=workbook["Sheet1"]                    #get sheet from workbook

# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="nagu.tcl@gmail.com"      #same data we write using loop statement


sheet.cell(1,1).value="nagu.tcl@gmail.com"              #multiple data write in excel sheet
sheet.cell(1,2).value="Nagarajan"

sheet.cell(2,1).value="Suganya.tcl@gmail.com"
sheet.cell(2,2).value="Suganya"

sheet.cell(3,1).value="Deeksha.tcl@gmail.com"
sheet.cell(3,2).value="Deeksha"

workbook.save(file)



















































